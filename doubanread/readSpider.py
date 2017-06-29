#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')


# Some User Agents
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

douban_read_root_url = 'https://read.douban.com'

def author_spider():
    page_num = 0
    book_list=[]
    try_times = 0

    while(1):
        url = 'https://read.douban.com/kind/0?start='+str(page_num*20)+'&sort=hot'
        time.sleep(np.random.rand()*5)

        try:
            req = urllib2.Request(url, headers=hds[page_num % len(hds)])
            source_code = urllib2.urlopen(req).read()
            plain_text = str(source_code)
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue

        soup = BeautifulSoup(plain_text,'html.parser')
        list_soup = soup.find('ul',{'class':'list-lined ebook-list column-list'})
        
        try_times += 1
        if list_soup == None and try_times < 200:
            continue
        elif list_soup == None or len(list_soup) <= 1:
            break  # Break when no informatoin got after 200 times requesting

        for book_info in list_soup.findAll('li'):
            book_item = book_info.find('div',{'class':'title'})
            if book_item:
                title = book_item.find('a').string.strip()
            else:
                book_h4=book_info.find('h4',{'class':'title'})
                if book_h4:
                    title = book_h4.find('a').string.strip()
                else:
                    continue
            author_item = book_info.find('a',{'class':'author-item'})
            book_type_item = book_info.find('span',{'class':'labeled-text'})
            book_type = ""
            if book_type_item:
                book_type_span = book_type_item.find('span')
                if book_type_span:
                    book_type = book_type_span.string.strip()   
            if author_item:
                author = author_item.string.strip()
                author_url = douban_read_root_url+author_item['href']
            else:
                author_div = book_info.find('div',{'class':'author'})
                if author_div:
                    author = author_div.find('a').string.strip()
                    author_url = douban_read_root_url+author_div.find('a')['href']
                else:
                    continue
            if not('author' in author_url):
                continue
            people_num = get_people_num(author_url)
            # book_score = book_info.find('span',{'class':'rating-average'}).string.strip()
            # book_comment_count = book_info.find('a',{'class':'ratings-link'}).find('span').string.strip()
            book_list.append([title,book_type,author,author_url,people_num])
        page_num +=1
        print 'Downloading Information From Page %d' % page_num
    return book_list

def get_people_num(url):
    try:
        time.sleep(1)
        req = urllib2.Request(url, headers=hds[np.random.randint(0, len(hds))])
        source_code = urllib2.urlopen(req).read()
        plain_text = str(source_code)
    except (urllib2.HTTPError, urllib2.URLError), e:
        print e
    soup = BeautifulSoup(plain_text,'html.parser')
    people_num = soup.find('div', {'class': 'followed-number'}).string.strip()
    return people_num


# 保存数据的Excel
def print_book_lists_excel(book_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    ws.append(wb.create_sheet(title='data'))
    ws[0].append(['序号', '书名','类别', '作者', '作者主页地址','粉丝数'])
    count = 1
    for bl in book_lists:
        ws[0].append([count, bl[0], bl[1],
                          bl[2],bl[3],bl[4]])
        count += 1
    save_path = 'book_list.xlsx'
    wb.save(save_path)
    print 'Downloading finished'



if __name__ == '__main__':
    book_list = author_spider()
    print_book_lists_excel(book_list)