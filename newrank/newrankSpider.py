#coding=utf-8
import sys
import time
import io
import json
import urllib
import urllib2
import hashlib
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from pyvirtualdisplay import Display


reload(sys)
sys.setdefaultencoding('utf8')


# Some User Agents
hds = [{'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},
       {'User-Agent': 'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},
       {'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

rank_name_group=['资讯','生活']
rank_name=[['时事','民生','财富','科技','创业','汽车','楼市','职场','教育','学术','政务','企业'],
['文化','百科','健康','时尚','美食','乐活','旅行','幽默','情感','体娱','美体','文摘']
]

from random import Random
#生成随机数
def random_str(randomlength=9):
    str = ''
    chars = 'abcdefghijklmnopqrstuvwxyz0123456789'
    length = len(chars) - 1
    random = Random()
    for i in range(randomlength):
        str+=chars[random.randint(0, length)]
    return str

#md5加密方法
def md5(str):
    import hashlib
    m = hashlib.md5()   
    m.update(str)
    return m.hexdigest()

#获取加密参数
def getsign(url,params):
    i = url+'?AppKey=joker&'
    randomStr = random_str()
    items = params.items()
    items.sort()
    for key,value in items:
        param = key+'='+value+'&'
        if key!='nonce' and key!='xyz':
            i+=param
    i += 'nonce='+ randomStr
    print i
    xyz = md5(i)
    params['nonce']=randomStr
    params['xyz']=xyz
    return params

#获取参数
def getParams():
    params={}
    params['end']='2017-05-31'
    params['start']='2017-05-01'
    return params

#获取排行榜列表
def getRankList(params):
   url = 'http://www.newrank.cn/xdnphb/list/month/rank' 
   datas = getsign('/xdnphb/list/month/rank',params)
   time.sleep(np.random.rand() * 5)
   try:
       data = urllib.urlencode(datas)
       request = urllib2.Request(url)
       opener = urllib2.build_opener(urllib2.HTTPCookieProcessor())  
       result = opener.open(request,data).read()
       #print params['rank_name']+'获得排行榜列表为：'+str(result)
       if result:
           return json.loads(result)
       else:
           return {'value':[]}
   except (urllib2.HTTPError, urllib2.URLError), e:
       print e
       return {'value':[]}

#启动firefox
def setup():
    browser = webdriver.Firefox()
    return browser

#模拟登陆
def login(username, password, browser=None):
    browser.get("http://www.newrank.cn/public/login/login.html")
    tab = browser.find_elements_by_class_name('login-normal-tap')[1]
    tab.click()
    pwd_btn = browser.find_element_by_id('password_input')
    act_btn = browser.find_element_by_id("account_input")
    submit_btn = browser.find_element_by_id("pwd_confirm")  

    act_btn.send_keys(username)
    pwd_btn.send_keys(password)
    submit_btn.click()
    time.sleep(5)
    return browser

#request加cookies
def set_session(browser):
    request = requests.Session()
    request.headers.update(hds[1])
    cookies = browser.get_cookies()
    for cookie in cookies:
        request.cookies.set(cookie['name'],cookie['value'])
    return request


#获取公众号详情页面
def getPeopleNum(request,account):

    url='http://www.newrank.cn/public/info/detail.html?account='+account
    try:
        print account +'正在查询中。。'
        response = request.get(url)
        soup = BeautifulSoup(response.text)
        item = soup.find('div',{'class':'detail-fans-counts'})
        if item:
            people_num = item['data'].replace(',','',5)
            print account+'的粉丝数为：'+people_num
            return people_num
        else:
            print account+'已下架'
            return 0
    except (urllib2.HTTPError, urllib2.URLError), e:
        print account+'抓取出错'
        return 0
    
# 保存数据的Excel
def print_lists_excel(wechat_lists):
    wb = Workbook(optimized_write=True)
    ws = []
    index = 0
    for key in wechat_lists:
        ws.append(wb.create_sheet(title=key.decode('utf-8')))
        ws[index].append(['序号','名称', '账号', '粉丝数'])
        count = 1
        for data in wechat_lists[key]:
            ws[index].append([count, data['name'], data['account'],
                          data['peopleNum']])
            count +=1
        index+=1
    save_path = '新榜.xlsx'
    wb.save(save_path)


if __name__ == '__main__':
    allResult = {}
    params = getParams()
    browser = login('18624382369','12345678',setup())
    rq = set_session(browser)
    for i,groupName in enumerate(rank_name_group):
        rankNameList = rank_name[i]
        params['rank_name_group']=groupName
        for rankName in rankNameList:
            params['rank_name']=rankName
            rankList = getRankList(params)['value']
            for wechat in rankList:
                peopleNum = getPeopleNum(rq,wechat['account'])
                wechat['peopleNum'] = peopleNum
            allResult[rankName] = rankList
    print_lists_excel(allResult)
 